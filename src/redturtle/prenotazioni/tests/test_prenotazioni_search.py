# -*- coding: utf-8 -*-
from copy import deepcopy
from datetime import date
from datetime import datetime
from datetime import timedelta
from dateutil import parser
from freezegun import freeze_time
from io import BytesIO
from plone import api
from plone.app.testing import setRoles
from plone.app.testing import SITE_OWNER_NAME
from plone.app.testing import SITE_OWNER_PASSWORD
from plone.app.testing import TEST_USER_ID
from plone.app.testing import TEST_USER_PASSWORD
from plone.app.textfield import RichTextValue
from plone.restapi.testing import RelativeSession
from plone.testing.zope import Browser
from redturtle.prenotazioni.testing import REDTURTLE_PRENOTAZIONI_API_FUNCTIONAL_TESTING

import openpyxl
import transaction
import unittest


class TestPrenotazioniSearch(unittest.TestCase):
    """Test the restapi search endpoint (<portal_url>/@bookings)"""

    from_date = datetime.min.isoformat()
    layer = REDTURTLE_PRENOTAZIONI_API_FUNCTIONAL_TESTING

    def setUp(self):
        self.app = self.layer["app"]
        self.portal = self.layer["portal"]
        self.portal_url = self.portal.absolute_url()
        self.testing_fiscal_code = "TESTINGFISCALCODE"
        self.testing_booking_date = parser.parse("2023-04-28 16:00:00")
        self.booking_expiration_date = parser.parse("2023-04-28 16:00:00") + timedelta(
            days=100
        )

        setRoles(self.portal, TEST_USER_ID, ["Manager"])

        self.api_session = RelativeSession(self.portal_url)
        self.api_session.headers.update({"Accept": "application/json"})
        self.api_session.auth = (SITE_OWNER_NAME, SITE_OWNER_PASSWORD)

        self.browser = Browser(self.layer["app"])

        self.folder_prenotazioni = api.content.create(
            container=self.portal,
            type="PrenotazioniFolder",
            title="Prenota foo",
            description="",
            daData=date.today(),
            gates=["Gate A"],
        )

        api.content.create(
            type="PrenotazioneType",
            title="Type A",
            duration=30,
            container=self.folder_prenotazioni,
            gates=["all"],
            requirements=RichTextValue(
                "You need to bring your own food", "text/plain", "text/html"
            ),
        )

        week_table = deepcopy(self.folder_prenotazioni.week_table)
        week_table[0]["morning_start"] = "0700"
        week_table[0]["morning_end"] = "1000"
        self.folder_prenotazioni.week_table = week_table

        self.folder_prenotazioni2 = api.content.create(
            container=self.portal,
            type="PrenotazioniFolder",
            title="Prenota bar",
            description="",
            daData=date.today(),
            gates=["Gate A"],
        )

        api.content.create(
            type="PrenotazioneType",
            title="Type A",
            duration=10,
            container=self.folder_prenotazioni2,
            gates=["all"],
        )
        api.content.create(
            type="PrenotazioneType",
            title="Type B",
            duration=20,
            container=self.folder_prenotazioni2,
            gates=["all"],
        )

        week_table = deepcopy(self.folder_prenotazioni2.week_table)
        week_table[0]["morning_start"] = "0700"
        week_table[0]["morning_end"] = "1000"
        self.folder_prenotazioni2.week_table = week_table

        year = api.content.create(
            container=self.folder_prenotazioni,
            type="PrenotazioniYear",
            title="Year",
        )
        week = api.content.create(container=year, type="PrenotazioniWeek", title="Week")
        self.day_folder = api.content.create(
            container=week, type="PrenotazioniDay", title="Day"
        )
        self.day_folder1 = api.content.create(
            container=week, type="PrenotazioniDay", title="Day"
        )

        year = api.content.create(
            container=self.folder_prenotazioni2,
            type="PrenotazioniYear",
            title="Year",
        )
        week = api.content.create(container=year, type="PrenotazioniWeek", title="Week")
        self.day_folder2 = api.content.create(
            container=week, type="PrenotazioniDay", title="Day"
        )

        self.prenotazione_fscode = api.content.create(
            container=self.day_folder,
            type="Prenotazione",
            title="Prenotazione fscode",
            booking_date=self.testing_booking_date - timedelta(days=2),
            booking_expiration_date=self.booking_expiration_date,
            fiscalcode=self.testing_fiscal_code,
        )
        self.prenotazione_no_fscode = api.content.create(
            container=self.day_folder,
            type="Prenotazione",
            booking_date=self.testing_booking_date - timedelta(days=2),
            booking_expiration_date=self.booking_expiration_date,
            title="Prenotazione no fscode",
        )
        self.prenotazione_datetime = api.content.create(
            container=self.day_folder,
            type="Prenotazione",
            title="Prenotazione datetime",
            booking_date=self.testing_booking_date,
            booking_expiration_date=self.booking_expiration_date,
            fiscalcode=self.testing_fiscal_code,
        )
        self.prenotazione_datetime_plus2 = api.content.create(
            container=self.day_folder1,
            type="Prenotazione",
            title="Prenotazione datetime plus 2",
            booking_date=self.testing_booking_date + timedelta(days=2),
            booking_expiration_date=self.booking_expiration_date,
            fiscalcode=self.testing_fiscal_code,
        )
        self.prenotazione_datetime_plus4 = api.content.create(
            container=self.day_folder2,
            type="Prenotazione",
            title="Prenotazione datetime plus 4",
            booking_date=self.testing_booking_date + timedelta(days=4),
            booking_expiration_date=self.booking_expiration_date,
            fiscalcode=self.testing_fiscal_code,
        )
        self.prenotazione_gateA = api.content.create(
            container=self.day_folder1,
            type="Prenotazione",
            title="Prenotazione gate A",
            booking_date=self.testing_booking_date + timedelta(days=8),
            booking_expiration_date=self.booking_expiration_date + timedelta(days=9),
            fiscalcode=self.testing_fiscal_code,
            gate="Gate A",
        )
        self.prenotazione_typeA = api.content.create(
            container=self.day_folder1,
            type="Prenotazione",
            title="Prenotazione type A",
            booking_date=self.testing_booking_date + timedelta(days=8),
            booking_expiration_date=self.booking_expiration_date + timedelta(days=9),
            fiscalcode=self.testing_fiscal_code,
            booking_type="Type A",
        )
        self.prenotazione_confirmed = api.content.create(
            container=self.day_folder1,
            type="Prenotazione",
            title="Prenotazione confirmed",
            booking_date=self.testing_booking_date + timedelta(days=8),
            booking_expiration_date=self.booking_expiration_date + timedelta(days=9),
            fiscalcode=self.testing_fiscal_code,
        )

        api.content.transition(self.prenotazione_confirmed, to_state="confirmed")

        transaction.commit()

    def tearDown(self):
        self.api_session.close()

    def test_view_permission(self):
        self.assertEqual(
            self.api_session.get(
                f"{self.portal.absolute_url()}/@bookings?from={self.from_date}"
            ).status_code,
            200,
        )

        setRoles(self.portal, TEST_USER_ID, [])

        self.api_session.auth = (TEST_USER_ID, TEST_USER_PASSWORD)

        self.assertEqual(
            self.api_session.get(f"{self.portal.absolute_url()}/@bookings").status_code,
            401,
        )

    def test_search_by_fiscalcode(self):
        result_uids = [
            i["booking_id"]
            for i in self.api_session.get(
                f"{self.portal.absolute_url()}/@bookings/{self.testing_fiscal_code}?from={self.from_date}"  # noqa: E501
            ).json()["items"]
        ]

        self.assertIn(self.prenotazione_fscode.UID(), result_uids)
        self.assertNotIn(self.prenotazione_no_fscode.UID(), result_uids)

    def test_fullobjects(self):
        items = self.api_session.get(
            f"{self.portal.absolute_url()}/@bookings?from={self.from_date}&fullobjects=1"
        ).json()["items"]
        self.assertIn(
            {
                "content-type": "text/plain",
                "data": "<p>You need to bring your own food</p>",
                "encoding": "utf-8",
            },  # noqa: E501
            [i["requirements"] for i in items],
        )

    def test_search_by_fiscalcode_case_insensitive(self):
        # ABCDEF12G34H567I -> AbCdEf12G34H567i
        camelcase_fiscalcode = "".join(
            [
                c.upper() if i % 2 == 0 else c.lower()
                for i, c in enumerate(self.testing_fiscal_code)
            ]
        )
        result_uids = [
            i["booking_id"]
            for i in self.api_session.get(
                f"{self.portal.absolute_url()}/@bookings/{camelcase_fiscalcode}"  # noqa: E501
            ).json()["items"]
        ]

        self.assertIn(self.prenotazione_fscode.UID(), result_uids)
        self.assertNotIn(self.prenotazione_no_fscode.UID(), result_uids)

    def test_search_by_fiscalcode_traverse(self):
        res = self.api_session.get(
            f"{self.portal.absolute_url()}/@bookings/{self.testing_fiscal_code}"
        )
        self.assertEqual(res.status_code, 200)
        ids = [i["booking_id"] for i in res.json()["items"]]
        self.assertIn(self.prenotazione_fscode.UID(), ids)
        self.assertNotIn(self.prenotazione_no_fscode.UID(), ids)

    def test_search_by_date(self):
        # test by start date
        res = self.api_session.get(
            f"{self.portal.absolute_url()}/@bookings?from={str(self.testing_booking_date + timedelta(days=1))}&fiscalcode={self.testing_fiscal_code}"
        )

        self.assertEqual(res.status_code, 200)

        ids = [i["booking_id"] for i in res.json()["items"]]
        self.assertNotIn(self.prenotazione_datetime.UID(), ids)
        self.assertIn(self.prenotazione_datetime_plus2.UID(), ids)

        # test by end date
        result_uids = [
            i["booking_id"]
            for i in self.api_session.get(
                f"{self.portal.absolute_url()}/@bookings?to={str(self.testing_booking_date + timedelta(days=3))}&fiscalcode={self.testing_fiscal_code}"
            ).json()["items"]
        ]

        self.assertIn(self.prenotazione_datetime_plus2.UID(), result_uids)
        self.assertNotIn(self.prenotazione_datetime_plus4.UID(), result_uids)

        # test btw strart and end date
        result_uids = [
            i["booking_id"]
            for i in self.api_session.get(
                f"{self.portal.absolute_url()}/@bookings?from={str(self.testing_booking_date + timedelta(days=1))}&to={str(self.testing_booking_date + timedelta(days=3))}&fiscalcode={self.testing_fiscal_code}"
            ).json()["items"]
        ]

        self.assertIn(self.prenotazione_datetime_plus2.UID(), result_uids)
        self.assertNotIn(self.prenotazione_datetime_plus4.UID(), result_uids)
        self.assertNotIn(self.prenotazione_datetime.UID(), result_uids)

    def test_search_by_modified_date(self):
        modified_booking = api.content.create(
            container=self.day_folder1,
            type="Prenotazione",
            title="Prenotazione",
            booking_date=self.testing_booking_date + timedelta(days=2),
            booking_expiration_date=self.booking_expiration_date,
            fiscalcode=self.testing_fiscal_code,
        )

        # change modification date
        self.prenotazione_datetime_plus2.setModificationDate(self.testing_booking_date)
        modification_date = self.testing_booking_date + timedelta(days=1)
        modified_booking.setModificationDate(modification_date + timedelta(hours=3))

        self.prenotazione_datetime_plus2.reindexObject(idxs=["modified"])
        modified_booking.reindexObject(idxs=["modified"])
        transaction.commit()

        # test by start and end date, return both
        result = self.api_session.get(
            f"{self.portal.absolute_url()}/@bookings?from={str(self.testing_booking_date + timedelta(days=1))}&to={str(self.testing_booking_date + timedelta(days=3))}&fiscalcode={self.testing_fiscal_code}"
        ).json()
        result_uids = [x["booking_id"] for x in result["items"]]

        self.assertEqual(result["items_total"], 2)
        self.assertIn(self.prenotazione_datetime_plus2.UID(), result_uids)
        self.assertIn(modified_booking.UID(), result_uids)

        # test by start and end date and modified, return only one
        result = self.api_session.get(
            f"{self.portal.absolute_url()}/@bookings?from={str(self.testing_booking_date + timedelta(days=1))}&to={str(self.testing_booking_date + timedelta(days=3))}&fiscalcode={self.testing_fiscal_code}&modified_after={modification_date.isoformat()}"
        ).json()
        result_uids = [x["booking_id"] for x in result["items"]]

        self.assertEqual(result["items_total"], 1)
        self.assertNotIn(self.prenotazione_datetime_plus2.UID(), result_uids)
        self.assertIn(modified_booking.UID(), result_uids)

    def test_search_inside_a_folder(self):
        res = self.api_session.get(
            f"{self.folder_prenotazioni2.absolute_url()}/@bookings?from={self.from_date}"
        )
        self.assertEqual(res.status_code, 200)
        self.assertEqual(res.json()["items_total"], 1)
        self.assertEqual(len(res.json()["items"]), 1)

    def test_search_by_gate(self):
        result_uids = [
            i["booking_id"]
            for i in self.api_session.get(
                f"{self.portal.absolute_url()}/@bookings?gate=Gate%20A"
            ).json()["items"]
        ]

        self.assertEqual(len(result_uids), 1)
        self.assertIn(self.prenotazione_gateA.UID(), result_uids)

    def test_search_by_booking_type(self):
        result_uids = [
            i["booking_id"]
            for i in self.api_session.get(
                f"{self.portal.absolute_url()}/@bookings?booking_type=Type+A"
            ).json()["items"]
        ]

        self.assertEqual(len(result_uids), 1)
        self.assertIn(self.prenotazione_typeA.UID(), result_uids)

    def test_search_by_review_state(self):
        result_uids = [
            i["booking_id"]
            for i in self.api_session.get(
                f"{self.portal.absolute_url()}/@bookings?review_state=confirmed"
            ).json()["items"]
        ]

        self.assertEqual(len(result_uids), 1)
        self.assertIn(self.prenotazione_confirmed.UID(), result_uids)

    def test_sort(self):
        res = self.api_session.get(
            f"{self.portal.absolute_url()}/@bookings?from={self.from_date}"
        )
        # default sort Date, descending
        self.assertEqual(
            [b["booking_date"] for b in res.json()["items"]],
            sorted([b["booking_date"] for b in res.json()["items"]], reverse=True),
        )
        res = self.api_session.get(
            f"{self.portal.absolute_url()}/@bookings?sort_order=ascending&from={self.from_date}"
        )
        self.assertEqual(
            [b["booking_date"] for b in res.json()["items"]],
            sorted([b["booking_date"] for b in res.json()["items"]]),
        )
        # sort on Title
        res = self.api_session.get(
            f"{self.portal.absolute_url()}/@bookings?sort_on=sortable_title&from={self.from_date}"
        )
        self.assertEqual(
            [b["title"] for b in res.json()["items"]],
            sorted([b["title"] for b in res.json()["items"]], reverse=True),
        )
        res = self.api_session.get(
            f"{self.portal.absolute_url()}/@bookings?sort_on=sortable_title&sort_order=ascending&from={self.from_date}"
        )
        self.assertEqual(
            [b["title"] for b in res.json()["items"]],
            sorted([b["title"] for b in res.json()["items"]]),
        )

    def test_download_xlsx(self):
        self.browser.addHeader(
            "Authorization",
            f"Basic {SITE_OWNER_NAME}:{SITE_OWNER_PASSWORD}",
        )
        self.browser.open(
            f"{self.folder_prenotazioni.absolute_url()}/@@download/bookings.xlsx"
        )
        self.assertEqual(self.browser._response.status, "200 OK")
        data = openpyxl.load_workbook(BytesIO(self.browser._response.body))
        self.assertEqual(len(list(data["Sheet 1"].rows)), 8)
        # TODO: test sorting
        # self.assertEqual(
        #     [r[1].value for r in data["Sheet 1"].rows],
        #     [
        #         "Stato",
        #         "Private",
        #         "Private",
        #         "Confermato",
        #         "Private",
        #         "Private",
        #         "Private",
        #         "Private",
        #     ],
        # )
        # BBB: sorting before check, because of the potential different order of the rows
        self.assertEqual(
            sorted([r[1].value for r in data["Sheet 1"].rows]),
            sorted(
                [
                    "Stato",
                    "Private",
                    "Private",
                    "Confermato",
                    "Private",
                    "Private",
                    "Private",
                    "Private",
                ]
            ),
        )

        self.browser.open(
            f"{self.folder_prenotazioni.absolute_url()}/@@download/bookings.xlsx?review_state=confirmed"
        )
        self.assertEqual(self.browser._response.status, "200 OK")
        data = openpyxl.load_workbook(BytesIO(self.browser._response.body))
        self.assertEqual(len(list(data["Sheet 1"].rows)), 2)
        self.assertEqual(
            [r[1].value for r in data["Sheet 1"].rows],
            ["Stato", "Confermato"],
        )

    def test_download_xlsx_post(self):
        self.browser.addHeader(
            "Authorization",
            f"Basic {SITE_OWNER_NAME}:{SITE_OWNER_PASSWORD}",
        )
        self.browser.open(
            f"{self.folder_prenotazioni.absolute_url()}/@@download/bookings.xlsx",
            data={
                "review_state": "confirmed",
            },
        )
        self.assertEqual(self.browser._response.status, "200 OK")
        data = openpyxl.load_workbook(BytesIO(self.browser._response.body))
        self.assertEqual(len(list(data["Sheet 1"].rows)), 2)
        self.assertEqual(
            [r[1].value for r in data["Sheet 1"].rows],
            ["Stato", "Confermato"],
        )

    def test_download_xlsx_sort(self):
        self.browser.addHeader(
            "Authorization",
            f"Basic {SITE_OWNER_NAME}:{SITE_OWNER_PASSWORD}",
        )
        self.browser.open(
            f"{self.folder_prenotazioni.absolute_url()}/@@download/bookings.xlsx?sort_order=ascending&sort_on=sortable_title"  # noqa: E501
        )
        self.assertEqual(self.browser._response.status, "200 OK")
        data = openpyxl.load_workbook(BytesIO(self.browser._response.body))
        self.assertEqual(len(list(data["Sheet 1"].rows)), 8)
        self.assertEqual(
            [r[0].value for r in data["Sheet 1"].rows][1:],
            sorted([r[0].value for r in data["Sheet 1"].rows][1:]),
        )

        self.browser.open(
            f"{self.folder_prenotazioni.absolute_url()}/@@download/bookings.xlsx?sort_order=descending&sort_on=sortable_title"  # noqa: E501
        )
        self.assertEqual(self.browser._response.status, "200 OK")
        data = openpyxl.load_workbook(BytesIO(self.browser._response.body))
        self.assertEqual(len(list(data["Sheet 1"].rows)), 8)
        self.assertEqual(
            [r[0].value for r in data["Sheet 1"].rows][1:],
            sorted([r[0].value for r in data["Sheet 1"].rows][1:], reverse=True),
        )

    def test_empty_query(self):
        res = self.api_session.get(
            f"{self.portal.absolute_url()}/@bookings?sort_order=descending&sort_on=sortable_title"
        )

        self.assertEqual(res.status_code, 400)


class TestPrenotazioniUserSearch(unittest.TestCase):
    """Test the restapi search endpoint (<portal_url>/@bookings)"""

    from_date = datetime.min.isoformat()
    layer = REDTURTLE_PRENOTAZIONI_API_FUNCTIONAL_TESTING

    def setUp(self):
        self.app = self.layer["app"]
        self.portal = self.layer["portal"]
        self.portal_url = self.portal.absolute_url()
        self.testing_fiscal_code = "TESTINGFISCALCODE"
        self.testing_booking_date = parser.parse("2023-04-28 16:00:00")
        self.booking_expiration_date = parser.parse("2023-04-28 16:00:00") + timedelta(
            days=100
        )

        setRoles(self.portal, TEST_USER_ID, ["Manager"])

        api.user.create(
            email="john@example.org",
            username="TESTINGFISCALCODE",
            password="testingpassowrd",
            properties={"fiscalcode": "TESTINGFISCALCODE"},
        )

        self.api_session = RelativeSession(self.portal_url)
        self.api_session.headers.update({"Accept": "application/json"})
        self.api_session.auth = (SITE_OWNER_NAME, SITE_OWNER_PASSWORD)

        self.anon_session = RelativeSession(self.portal_url)
        self.anon_session.headers.update({"Accept": "application/json"})
        self.anon_session.auth = None

        self.user_session = RelativeSession(self.portal_url)
        self.user_session.headers.update({"Accept": "application/json"})
        self.user_session.auth = ("TESTINGFISCALCODE", "testingpassowrd")

        self.browser = Browser(self.layer["app"])

        self.folder_prenotazioni = api.content.create(
            container=self.portal,
            type="PrenotazioniFolder",
            title="Prenota foo",
            description="",
            daData=parser.parse("2023-04-01").date(),
            gates=["Gate A"],
        )
        api.content.transition(obj=self.folder_prenotazioni, transition="publish")

        obj = api.content.create(
            type="PrenotazioneType",
            title="Type A",
            duration=30,
            container=self.folder_prenotazioni,
            gates=["all"],
            requirements=RichTextValue(
                "You need to bring your own food", "text/plain", "text/html"
            ),
            booking_additional_fields_schema=[
                {
                    "name": "foo",
                    "label": "This is Foo",
                    "description": "text field description",
                    "type": "text",
                }
            ],
        )
        api.content.transition(obj=obj, transition="publish")

        week_table = deepcopy(self.folder_prenotazioni.week_table)
        week_table[0]["morning_start"] = "0700"
        week_table[0]["morning_end"] = "1000"
        self.folder_prenotazioni.week_table = week_table

        self.folder_prenotazioni2 = api.content.create(
            container=self.portal,
            type="PrenotazioniFolder",
            title="Prenota bar",
            description="",
            daData=date.today(),
            gates=["Gate A"],
        )
        api.content.transition(obj=self.folder_prenotazioni2, transition="publish")

        obj = api.content.create(
            type="PrenotazioneType",
            title="Type A",
            duration=10,
            container=self.folder_prenotazioni2,
            gates=["all"],
        )
        api.content.transition(obj=obj, transition="publish")

        obj = api.content.create(
            type="PrenotazioneType",
            title="Type B",
            duration=20,
            container=self.folder_prenotazioni2,
            gates=["all"],
        )
        api.content.transition(obj=obj, transition="publish")

        week_table = deepcopy(self.folder_prenotazioni2.week_table)
        week_table[0]["morning_start"] = "0700"
        week_table[0]["morning_end"] = "1000"
        self.folder_prenotazioni2.week_table = week_table

        transaction.commit()

    def tearDown(self):
        self.api_session.close()
        self.anon_session.close()
        self.user_session.close()

    # utility methods
    def add_booking(
        self, api_session, booking_date, booking_type, fields, additional_fields=None
    ):
        return api_session.post(
            f"{self.folder_prenotazioni.absolute_url()}/@booking",
            json={
                "booking_date": booking_date,
                "booking_type": booking_type,
                "fields": fields,
                "additional_fields": additional_fields,
            },
        )

    @freeze_time("2023-05-14")
    def test_search_own_bookings(self):
        # booking_date = "{}T09:00:00+00:00".format(
        #     (date.today() + timedelta(1)).strftime("%Y-%m-%d")
        # )
        res = self.anon_session.get(
            f"{self.folder_prenotazioni.absolute_url()}/@available-slots"
        )

        booking_date = res.json()["items"][0]
        # anonymous user 1
        res = self.add_booking(
            self.anon_session,
            booking_date=booking_date,
            booking_type="Type A",
            fields=[
                {"name": "title", "value": "Mario Rossi"},
                {"name": "email", "value": "mario.rossi@example"},
                {"name": "fiscalcode", "value": "ABCDEF12G34H567I"},
            ],
        )
        self.assertEqual(res.status_code, 200)
        # anonymous user 2 (not the same fiscalcode, pretend to be john)
        res = self.anon_session.get(
            f"{self.folder_prenotazioni.absolute_url()}/@available-slots"
        )
        booking_date = res.json()["items"][0]
        res = self.add_booking(
            self.anon_session,
            booking_date=booking_date,
            booking_type="Type A",
            fields=[
                {"name": "title", "value": "John Rossi"},
                {"name": "email", "value": "john@example"},
                {"name": "fiscalcode", "value": "TESTINGFISCALCODE"},
            ],
        )
        self.assertEqual(res.status_code, 200)
        # john
        res = self.anon_session.get(
            f"{self.folder_prenotazioni.absolute_url()}/@available-slots"
        )
        booking_date = res.json()["items"][0]
        res = self.add_booking(
            self.user_session,
            booking_date=booking_date,
            booking_type="Type A",
            fields=[
                {"name": "title", "value": "John Rossi"},
                {"name": "email", "value": "john@example"},
                # {"name": "fiscalcode", "value": "TESTINGFISCALCODE"},
            ],
        )
        self.assertEqual(res.status_code, 200)

        # anonimo non può vedere le prenotazioni
        res = self.anon_session.get(
            f"{self.portal.absolute_url()}/@bookings/TESTINGFISCALCODE"
        )
        self.assertEqual(res.status_code, 401)

        # il manager vede tutte le prenotazioni (3)
        res = self.api_session.get(
            f"{self.portal.absolute_url()}/@bookings?from={self.from_date}"
        )
        self.assertEqual(res.status_code, 200)
        self.assertEqual(res.json()["items_total"], 3)

        # la prenotazione fatta da anonimo con il codicefiscale di john non è visibile
        res = self.user_session.get(
            f"{self.portal.absolute_url()}/@bookings?from={self.from_date}"
        )
        self.assertEqual(res.status_code, 200)
        self.assertEqual(res.json()["items_total"], 1)

        # XXX: monkeypatching per vedere le prenotazioni anonime
        from redturtle.prenotazioni.restapi.services.bookings import search

        search.SEE_OWN_ANONYMOUS_BOOKINGS = True
        res = self.user_session.get(
            f"{self.portal.absolute_url()}/@bookings?from={self.from_date}"
        )
        self.assertEqual(res.status_code, 200)
        self.assertEqual(res.json()["items_total"], 2)
        search.SEE_OWN_ANONYMOUS_BOOKINGS = False

        # TODO: verificare che siano le prenotazioni giuste

    @freeze_time("2023-05-14")
    def test_additional_fields(self):
        res = self.anon_session.get(
            f"{self.folder_prenotazioni.absolute_url()}/@available-slots"
        )
        booking_date = res.json()["items"][0]
        res = self.add_booking(
            self.anon_session,
            booking_date=booking_date,
            booking_type="Type A",
            fields=[
                {"name": "title", "value": "Mario Rossi"},
                {"name": "email", "value": "mario.rossi@example"},
                {"name": "fiscalcode", "value": "ABCDEF12G34H567I"},
            ],
            additional_fields=[{"name": "foo", "value": "bar"}],
        )
        self.assertEqual(res.status_code, 200)
        self.assertEqual(
            res.json()["additional_fields"], [{"name": "foo", "value": "bar"}]
        )
        booking_code = res.json()["booking_code"]
        res = self.api_session.get(
            f"{self.portal.absolute_url()}/@bookings?SearchableText={booking_code}"
        )
        self.assertEqual(res.status_code, 200)
        self.assertEqual(res.json()["items_total"], 1)
        self.assertEqual(
            res.json()["items"][0]["additional_fields"],
            [{"name": "foo", "value": "bar"}],
        )
